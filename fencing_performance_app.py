import streamlit as st
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings('ignore')
from datetime import datetime
import openpyxl
from io import BytesIO

# レポート生成用ライブラリ
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_pdf import PdfPages
import seaborn as sns
import base64

# Plotlyが利用可能かチェック
try:
    import plotly.express as px
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    st.warning("Plotly library not found. Graph functionality will be disabled.")

# レポート用の変数設定
REPORT_METRICS = {
    'CMJ': ['Jump Height(cm)', 'mRSI', 'Braking RFD'],
    'IMTP': ['Relative Peak Force (BW)']
}

# ページ設定
st.set_page_config(
    page_title="Fencing Performance Test",
    page_icon="🔲",
    layout="wide",
    initial_sidebar_state="expanded"
)

# カスタムCSS（シックなデザイン）
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #2D3748 0%, #1A202C 100%);
        padding: 2.5rem;
        border-radius: 0px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
        font-weight: 700;
        font-size: 2.8rem;
        box-shadow: 0 8px 32px rgba(45, 55, 72, 0.25);
        border-left: 6px solid #171923;
    }
    
    .section-header {
        background: linear-gradient(135deg, #4A5568 0%, #2D3748 100%);
        padding: 1.2rem 2rem;
        border-radius: 0px;
        color: white;
        font-weight: 600;
        margin: 2rem 0 1.5rem 0;
        font-size: 1.4rem;
        box-shadow: 0 4px 16px rgba(74, 85, 104, 0.2);
        border-left: 4px solid #1A202C;
    }
    
    .metric-card {
        background: linear-gradient(135deg, #718096 0%, #4A5568 100%);
        padding: 2rem;
        border-radius: 0px;
        margin: 0.75rem;
        color: white;
        text-align: center;
        box-shadow: 0 8px 24px rgba(113, 128, 150, 0.15);
        transition: all 0.3s ease;
        border: 1px solid rgba(255, 255, 255, 0.1);
    }
    
    .metric-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 12px 32px rgba(113, 128, 150, 0.25);
    }
    
    .highlight-metric {
        font-size: 2.4rem;
        font-weight: 700;
        margin: 0.8rem 0;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
    }
    
    .metric-label {
        font-size: 1.2rem;
        margin-bottom: 0.8rem;
        opacity: 0.95;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    .comparison-text {
        font-size: 1rem;
        opacity: 0.85;
        margin-top: 0.8rem;
        font-weight: 400;
    }
    
    .stDataFrame {
        background: white;
        border-radius: 8px;
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.08);
        overflow: hidden;
    }
    
    .player-title {
        color: #2D3748;
        font-size: 2.2rem;
        font-weight: 700;
        margin-bottom: 1rem;
        padding: 1rem 0;
        border-bottom: 3px solid #718096;
    }
    
    .date-info {
        background: linear-gradient(135deg, #F7FAFC 0%, #E2E8F0 100%);
        padding: 1rem;
        border-radius: 8px;
        color: #2D3748;
        font-weight: 500;
        text-align: center;
        border: 1px solid #CBD5E0;
    }
    
    .page-nav {
        background: linear-gradient(135deg, #E2E8F0 0%, #CBD5E0 100%);
        padding: 1rem;
        border-radius: 8px;
        margin-bottom: 2rem;
        text-align: center;
    }
    
    .report-section {
        background: linear-gradient(135deg, #EDF2F7 0%, #E2E8F0 100%);
        padding: 2rem;
        border-radius: 8px;
        margin: 2rem 0;
        border-left: 4px solid #2D3748;
    }
</style>
""", unsafe_allow_html=True)

def sheet_to_dataframe(sheet):
    """シートをDataFrameに変換"""
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # 空行をスキップ
            data.append(list(row))
    
    if not data or len(data) < 2:
        return pd.DataFrame()
        
    # ヘッダー行を取得
    headers = data[0]
    
    # データ行を取得
    rows = data[1:]
    
    # DataFrameを作成
    df_data = {}
    for i, header in enumerate(headers):
        if header is not None:
            column_data = []
            for row in rows:
                if i < len(row):
                    value = row[i]
                    # 日付列の特別処理
                    if str(header).lower() == 'date' and value is not None:
                        try:
                            # Excelのdatetimeオブジェクトを確実にPythonのdatetimeに変換
                            if hasattr(value, 'date'):
                                # すでにdatetimeオブジェクトの場合
                                column_data.append(pd.Timestamp(value))
                            elif isinstance(value, (int, float)):
                                # Excelのシリアル番号の場合
                                column_data.append(pd.Timestamp('1900-01-01') + pd.Timedelta(days=value-2))
                            else:
                                # その他の場合は文字列として解析
                                column_data.append(pd.to_datetime(str(value)))
                        except Exception as e:
                            column_data.append(None)
                    else:
                        column_data.append(value)
                else:
                    column_data.append(None)
            df_data[str(header)] = column_data
    
    # pandasのDataFrameに変換
    df = pd.DataFrame(df_data)
    
    return df

def load_excel_manually(uploaded_file):
    """手動でExcelファイルを読み込む"""
    try:
        # ファイルをバイト形式で読み込み
        file_content = uploaded_file.getvalue()
        
        # openpyxlでワークブックを開く
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        
        # シート名を確認
        if 'CMJ' not in wb.sheetnames or 'IMTP' not in wb.sheetnames:
            st.error(f"Required sheets not found. Available sheets: {wb.sheetnames}")
            return None
        
        # 各シートからデータを取得
        data_dict = {}
        
        for sheet_name in ['CMJ', 'IMTP']:
            sheet = wb[sheet_name]
            df = sheet_to_dataframe(sheet)
            df['Type'] = sheet_name
            data_dict[sheet_name] = df
        
        return data_dict
        
    except Exception as e:
        st.error(f"Manual Excel loading error: {str(e)}")
        return None

def create_dataframe_from_dict(data_dict):
    """辞書からDataFrameを作成"""
    try:
        dfs = []
        
        for sheet_name, df in data_dict.items():
            # 空行を除去
            df = df.dropna(subset=['Name'])
            
            # 日付列を確実にdatetime型に変換
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
            
            dfs.append(df)
        
        if dfs:
            # 結合
            combined_df = pd.concat(dfs, ignore_index=True, sort=False)
            
            # デバッグ情報：日付の範囲を表示
            if 'Date' in combined_df.columns:
                date_range = combined_df['Date'].dropna()
                if not date_range.empty:
                    st.success(f"✅ Data loaded! Date range: {date_range.min().strftime('%Y-%m-%d')} to {date_range.max().strftime('%Y-%m-%d')}")
            
            return combined_df
        else:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"DataFrame creation error: {str(e)}")
        return pd.DataFrame()

def get_test_config():
    """Test configuration"""
    return {
        'CMJ': {
            'name': 'Counter Movement Jump',
            'metrics': [
                'Jump Height(cm)',
                'Countermovement Depth', 
                'Braking RFD',
                'Avg. Braking Force',
                'Avg. Propulsive Force',
                'mRSI'
            ],
            'units': {
                'Jump Height(cm)': 'cm',
                'Countermovement Depth': 'm',
                'Braking RFD': 'N/s',
                'Avg. Braking Force': 'N',
                'Avg. Propulsive Force': 'N',
                'mRSI': ''
            },
            'highlight': ['Jump Height(cm)', 'mRSI', 'Avg. Propulsive Force'],
            'female_norms': {
                'Jump Height(cm)': {'mean': 33.65, 'std': 4.28},
                'mRSI': {'mean': 0.47, 'std': 0.08},
                'Braking RFD': {'mean': 6594.37, 'std': 1858.18}
            }
        },
        'IMTP': {
            'name': 'Isometric Mid-Thigh Pull',
            'metrics': [
                'Peak Force',
                'Relative Peak Force (BW)',
                'RFD 0-50 ms',
                'RFD 0-100 ms',
                'RFD 0-150 ms',
                'RFD 0-200 ms',
                'RFD 0-250 ms'
            ],
            'units': {
                'Peak Force': 'N',
                'Relative Peak Force (BW)': 'BW',
                'RFD 0-50 ms': 'N/s',
                'RFD 0-100 ms': 'N/s',
                'RFD 0-150 ms': 'N/s',
                'RFD 0-200 ms': 'N/s',
                'RFD 0-250 ms': 'N/s'
            },
            'highlight': ['Peak Force', 'Relative Peak Force (BW)', 'RFD 0-100 ms'],
            'female_norms': {
                'Relative Peak Force (BW)': {'mean': 42.45, 'std': 7.21},
                'RFD 0-250 ms': {'mean': 102.43, 'std': 23.89}
            }
        }
    }

def safe_get_value(data, column, default=None):
    """安全に最新値を取得"""
    try:
        if column not in data.columns or data.empty:
            return default
        
        valid_data = data[data[column].notna()]
        valid_data = valid_data[valid_data[column] != '']
        valid_data = valid_data[valid_data[column] != 0]
        
        if valid_data.empty:
            return default
        
        if 'Date' in valid_data.columns:
            # 日付を確実にdatetime型に変換してから降順ソート
            valid_data = valid_data.copy()
            valid_data['Date'] = pd.to_datetime(valid_data['Date'])
            latest_valid = valid_data.sort_values('Date', ascending=False).iloc[0]
            value = latest_valid[column]
        else:
            value = valid_data.iloc[0][column]
        
        if pd.isna(value) or value == '' or value == 0:
            return default
        
        if isinstance(value, (int, float, np.number)):
            if np.isfinite(value):
                return float(value)
        
        return default
        
    except:
        return default

def safe_get_best_value(data, column, default=None):
    """安全に最高値を取得"""
    try:
        if column not in data.columns or data.empty:
            return default, default
        
        valid_data = data[data[column].notna()]
        valid_data = valid_data[valid_data[column] != '']
        valid_data = valid_data[valid_data[column] != 0]
        
        if valid_data.empty:
            return default, default
        
        numeric_values = pd.to_numeric(valid_data[column], errors='coerce')
        clean_values = numeric_values.dropna()
        
        if clean_values.empty:
            return default, default
        
        max_value = clean_values.max()
        max_idx = clean_values.idxmax()
        
        best_date = "N/A"
        if 'Date' in data.columns and max_idx in data.index:
            date_val = data.loc[max_idx, 'Date']
            if pd.notna(date_val):
                # 日付を確実にdatetime型に変換
                date_val = pd.to_datetime(date_val)
                best_date = date_val.strftime('%Y-%m-%d')
        
        return float(max_value), best_date
        
    except:
        return default, default

def safe_mean(series):
    """安全に平均値を計算"""
    if series.empty:
        return None
    numeric_series = pd.to_numeric(series, errors='coerce')
    clean_series = numeric_series.dropna()
    clean_series = clean_series[clean_series != 0]
    return clean_series.mean() if len(clean_series) > 0 else None

def format_value(value, unit=""):
    """値を安全にフォーマット"""
    if value is None or pd.isna(value):
        return "N/A"
    try:
        formatted_val = f"{float(value):.2f}"
        return f"{formatted_val}{unit}" if unit else formatted_val
    except:
        return "N/A"

# レポート生成関数群
def create_individual_report(player_data, all_data, player_name):
    """個人レポートを作成"""
    plt.style.use('default')
    sns.set_palette("husl")
    
    # フィギュアサイズをA4に設定 (8.27 x 11.69 inch)
    fig = plt.figure(figsize=(8.27, 11.69))
    
    # グリッドレイアウト設定: 6行2列（Team部分を下に移動）
    gs = fig.add_gridspec(6, 2, height_ratios=[0.8, 1, 1, 0.8, 1, 1], hspace=0.7, wspace=0.35)
    
    # タイトル（さらに20ポイント下げる）
    fig.suptitle(f'Performance Report - {player_name}', fontsize=16, fontweight='bold', y=0.955)
    
    # 1. 個人の表（Team表と同じ高さに調整）
    ax_table1 = fig.add_subplot(gs[0, :])
    ax_table1.axis('off')
    ax_table1.text(0.5, 0.95, 'Individual Performance Summary', 
                   fontsize=14, fontweight='bold', ha='center', transform=ax_table1.transAxes)
    
    # 個人データテーブル作成
    individual_table_data = create_individual_summary_table(player_data)
    if individual_table_data:
        table1 = ax_table1.table(cellText=individual_table_data['data'],
                                colLabels=individual_table_data['headers'],
                                cellLoc='center',
                                loc='center',
                                bbox=[0.03, 0.15, 0.94, 0.75])  # 幅を20ポイント拡大 (0.05→0.03, 0.9→0.94)
        table1.auto_set_font_size(False)
        table1.set_fontsize(7)
        table1.scale(1, 2.2)  # Team表と同じ縦幅スケール
        
        # ヘッダーのスタイル設定
        for i in range(len(individual_table_data['headers'])):
            table1[(0, i)].set_facecolor('#2D3748')
            table1[(0, i)].set_text_props(weight='bold', color='white')
    
    # 2. 個人の4つの推移グラフ（サイズを統一）
    metric_list = ['Jump Height(cm)', 'mRSI', 'Braking RFD', 'Relative Peak Force (BW)']
    positions = [(1, 0), (1, 1), (2, 0), (2, 1)]  # 2x2グリッド
    
    for i, (metric, pos) in enumerate(zip(metric_list, positions)):
        ax = fig.add_subplot(gs[pos])
        create_single_metric_graph(ax, player_data, metric, f'Individual {metric}', individual=True)
    
    # 3. チーム比較の表
    ax_table2 = fig.add_subplot(gs[3, :])
    ax_table2.axis('off')
    ax_table2.text(0.5, 0.98, 'Team Average Comparison', 
                   fontsize=14, fontweight='bold', ha='center', transform=ax_table2.transAxes)
    
    # チーム比較テーブル作成
    team_table_data = create_team_comparison_summary_table(player_data, all_data)
    if team_table_data:
        table2 = ax_table2.table(cellText=team_table_data['data'],
                                colLabels=team_table_data['headers'],
                                cellLoc='center',
                                loc='center',
                                bbox=[0.03, 0.15, 0.94, 0.75])  # 幅を20ポイント拡大 (0.05→0.03, 0.9→0.94)
        table2.auto_set_font_size(False)
        table2.set_fontsize(7)
        table2.scale(1, 2.2)  # Individual表と同じ縦幅スケール
        
        # ヘッダーのスタイル設定
        for i in range(len(team_table_data['headers'])):
            table2[(0, i)].set_facecolor('#2D3748')
            table2[(0, i)].set_text_props(weight='bold', color='white')
    
    # 4. チームの4つの推移グラフ（サイズを統一）
    team_positions = [(4, 0), (4, 1), (5, 0), (5, 1)]  # 2x2グリッド
    
    for i, (metric, pos) in enumerate(zip(metric_list, team_positions)):
        ax = fig.add_subplot(gs[pos])
        create_single_metric_graph(ax, all_data, metric, f'Team {metric}', individual=False)
    
    # 日付情報を追加
    all_dates = player_data['Date'].dropna()
    if not all_dates.empty:
        all_dates = pd.to_datetime(all_dates).sort_values(ascending=False)
        latest_date = all_dates.iloc[0].strftime('%Y-%m-%d')
        oldest_date = all_dates.iloc[-1].strftime('%Y-%m-%d')
        fig.text(0.02, 0.005, f'Report Period: {oldest_date} to {latest_date}', 
                fontsize=8, ha='left')
    
    fig.text(0.98, 0.005, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', 
            fontsize=8, ha='right')
    
    plt.tight_layout()
    return fig

def create_individual_summary_table(player_data):
    """個人サマリーテーブルを作成"""
    headers = ['Metric', 'Latest Value', 'Personal Best', 'Test Date']
    data = []
    
    for test_type, metrics in REPORT_METRICS.items():
        test_data = player_data[player_data['Type'] == test_type]
        if test_data.empty:
            continue
            
        for metric in metrics:
            if metric not in test_data.columns:
                continue
                
            latest_val = safe_get_value(test_data, metric)
            best_val, best_date = safe_get_best_value(test_data, metric)
            
            # 最新のテスト日付取得
            latest_date = "N/A"
            if latest_val is not None:
                valid_data = test_data.dropna(subset=[metric])
                valid_data = valid_data[valid_data[metric] != 0]
                if not valid_data.empty and 'Date' in valid_data.columns:
                    valid_data = valid_data.copy()
                    valid_data['Date'] = pd.to_datetime(valid_data['Date'])
                    latest_valid = valid_data.sort_values('Date', ascending=False).iloc[0]
                    latest_date = latest_valid['Date'].strftime('%Y-%m-%d') if pd.notna(latest_valid['Date']) else "N/A"
            
            data.append([
                metric,
                format_value(latest_val),
                format_value(best_val),
                latest_date
            ])
    
    return {'headers': headers, 'data': data} if data else None

def create_team_comparison_summary_table(player_data, all_data):
    """チーム比較サマリーテーブルを作成"""
    headers = ['Metric', 'Individual', 'Team Average', 'Percentile Rank']
    data = []
    
    for test_type, metrics in REPORT_METRICS.items():
        test_data = player_data[player_data['Type'] == test_type]
        team_data = all_data[all_data['Type'] == test_type]
        
        if test_data.empty or team_data.empty:
            continue
            
        for metric in metrics:
            if metric not in test_data.columns or metric not in team_data.columns:
                continue
                
            player_val = safe_get_value(test_data, metric)
            team_avg = safe_mean(team_data[metric])
            
            # パーセンタイル計算
            percentile = "N/A"
            if player_val is not None:
                team_values = pd.to_numeric(team_data[metric], errors='coerce').dropna()
                team_values = team_values[team_values != 0]
                if len(team_values) > 0:
                    percentile_val = (team_values < player_val).sum() / len(team_values) * 100
                    percentile = f"{percentile_val:.0f}%"
            
            data.append([
                metric,
                format_value(player_val),
                format_value(team_avg),
                percentile
            ])
    
    return {'headers': headers, 'data': data} if data else None

def create_single_metric_graph(ax, data, metric, title, individual=True):
    """単一のメトリクスグラフを作成"""
    ax.set_title(title, fontsize=9, fontweight='bold', pad=6)
    
    colors = {'Jump Height(cm)': '#2D3748', 'mRSI': '#DC2626', 
              'Braking RFD': '#059669', 'Relative Peak Force (BW)': '#7C3AED'}
    color = colors.get(metric, '#2D3748')
    
    has_data = False
    
    if individual:
        # 個人データの処理
        for test_type, metrics in REPORT_METRICS.items():
            if metric in metrics:
                test_data = data[data['Type'] == test_type]
                if test_data.empty:
                    continue
                    
                if metric not in test_data.columns:
                    continue
                    
                # 有効なデータをフィルター
                valid_data = test_data.dropna(subset=[metric, 'Date'])
                valid_data = valid_data[valid_data[metric] != 0]
                
                if len(valid_data) < 1:
                    continue
                    
                # 日付変換とソート
                valid_data = valid_data.copy()
                valid_data['Date'] = pd.to_datetime(valid_data['Date'])
                valid_data = valid_data.sort_values('Date')
                
                # 数値変換
                valid_data[metric] = pd.to_numeric(valid_data[metric], errors='coerce')
                valid_data = valid_data.dropna(subset=[metric])
                
                if len(valid_data) < 1:
                    continue
                
                # プロット
                ax.plot(valid_data['Date'], valid_data[metric], 
                       marker='o', linewidth=2.5, markersize=5,
                       color=color, markerfacecolor='white', 
                       markeredgecolor=color, markeredgewidth=2)
                
                has_data = True
                break
    else:
        # チームデータの処理（月別平均）
        for test_type, metrics in REPORT_METRICS.items():
            if metric in metrics:
                test_data = data[data['Type'] == test_type]
                if test_data.empty:
                    continue
                    
                if metric not in test_data.columns:
                    continue
                    
                # 有効なデータをフィルター
                valid_data = test_data.dropna(subset=[metric, 'Date'])
                valid_data = valid_data[valid_data[metric] != 0]
                
                if len(valid_data) < 1:
                    continue
                
                # 日付変換
                valid_data = valid_data.copy()
                valid_data['Date'] = pd.to_datetime(valid_data['Date'])
                valid_data[metric] = pd.to_numeric(valid_data[metric], errors='coerce')
                valid_data = valid_data.dropna(subset=[metric])
                
                if len(valid_data) < 1:
                    continue
                
                # 月別平均を計算
                valid_data['YearMonth'] = valid_data['Date'].dt.to_period('M')
                monthly_avg = valid_data.groupby('YearMonth')[metric].mean().reset_index()
                monthly_avg['Date'] = monthly_avg['YearMonth'].dt.to_timestamp()
                
                if len(monthly_avg) < 1:
                    continue
                
                # プロット
                ax.plot(monthly_avg['Date'], monthly_avg[metric], 
                       marker='s', linewidth=2.5, markersize=5,
                       color=color, linestyle='--', alpha=0.8,
                       markerfacecolor='white', markeredgecolor=color, markeredgewidth=2)
                
                has_data = True
                break
    
    if has_data:
        ax.grid(True, alpha=0.3, linewidth=0.5)
        ax.tick_params(axis='x', rotation=45, labelsize=6)  # X軸文字サイズを6に縮小
        ax.tick_params(axis='y', labelsize=7)
        
        # X軸の日付フォーマット（より短い形式）
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))  # 3ヶ月間隔に変更
        
        # Y軸ラベル
        units = {
            'Jump Height(cm)': 'cm',
            'mRSI': '',
            'Braking RFD': 'N/s',
            'Relative Peak Force (BW)': 'BW'
        }
        unit = units.get(metric, '')
        if unit:
            ax.set_ylabel(unit, fontsize=7)
    else:
        ax.text(0.5, 0.5, 'No data available', 
               ha='center', va='center', transform=ax.transAxes,
               fontsize=8, color='gray')
        ax.set_xticks([])
        ax.set_yticks([])

def generate_pdf_report(player_data, all_data, player_name):
    """PDFレポートを生成してダウンロード可能な形式で返す"""
    # レポート作成
    fig = create_individual_report(player_data, all_data, player_name)
    
    # PDFに保存
    buffer = BytesIO()
    with PdfPages(buffer) as pdf:
        pdf.savefig(fig, bbox_inches='tight', dpi=300)
    
    plt.close(fig)
    buffer.seek(0)
    
    return buffer.getvalue()

def create_comparison_table(player_data, all_data, metrics, test_type, config):
    """比較テーブルを作成"""
    table_data = []
    
    test_data = all_data[all_data['Type'] == test_type]
    female_norms = config[test_type].get('female_norms', {})
    
    for metric in metrics:
        player_val = safe_get_value(player_data, metric)
        best_val, best_date = safe_get_best_value(player_data, metric)
        avg_val = safe_mean(test_data[metric])
        
        female_norm_text = "N/A"
        if metric in female_norms:
            mean_val = female_norms[metric]['mean']
            std_val = female_norms[metric]['std']
            female_norm_text = f"{mean_val:.2f} ± {std_val:.2f}"
        
        measurement_date = "N/A"
        if player_val is not None:
            valid_data = player_data.dropna(subset=[metric])
            valid_data = valid_data[valid_data[metric] != 0]
            if not valid_data.empty and 'Date' in valid_data.columns:
                # 日付を確実にdatetime型に変換してから降順ソート
                valid_data = valid_data.copy()
                valid_data['Date'] = pd.to_datetime(valid_data['Date'])
                latest_valid = valid_data.sort_values('Date', ascending=False).iloc[0]
                measurement_date = latest_valid['Date'].strftime('%Y-%m-%d') if pd.notna(latest_valid['Date']) else "N/A"
        
        best_value_text = "N/A"
        if best_val is not None:
            best_value_text = f"{best_val:.2f}"
            if best_date != "N/A":
                best_value_text += f" ({best_date})"
        
        table_data.append({
            'Metric': metric,
            'Latest Value': format_value(player_val),
            'Test Date': measurement_date,
            'Personal Best': best_value_text,
            'Team Average': format_value(avg_val),
            'Female Fencer Norm': female_norm_text
        })
    
    return pd.DataFrame(table_data)

def create_team_comparison_chart(df, selected_athletes, test_type, config):
    """複数選手の比較チャートを作成"""
    if not PLOTLY_AVAILABLE:
        return None
    
    if not selected_athletes:
        return None
    
    test_config = config[test_type]
    metrics = test_config['metrics']
    units = test_config['units']
    
    # 選手ごとの色を定義（選手数に応じて色を調整）
    athlete_colors = [
        '#2D3748', '#DC2626', '#059669', '#7C3AED', '#EA580C', 
        '#0891B2', '#BE185D', '#65A30D', '#9333EA', '#C2410C'
    ]
    
    # データをフィルター
    team_data = df[(df['Type'] == test_type) & (df['Name'].isin(selected_athletes))].copy()
    
    if team_data.empty:
        return None
    
    # 日付でソート
    team_data['Date'] = pd.to_datetime(team_data['Date'])
    team_data = team_data.sort_values('Date')
    
    # サブプロットの設定
    rows = (len(metrics) + 1) // 2
    cols = min(2, len(metrics))
    
    fig = make_subplots(
        rows=rows,
        cols=cols,
        subplot_titles=[f"<b>{metric}</b>" for metric in metrics],
        vertical_spacing=0.18,
        horizontal_spacing=0.15
    )
    
    for i, metric in enumerate(metrics):
        if metric not in team_data.columns:
            continue
            
        row = (i // 2) + 1
        col = (i % 2) + 1
        
        for j, athlete in enumerate(selected_athletes):
            athlete_data = team_data[team_data['Name'] == athlete]
            
            if athlete_data.empty:
                continue
            
            # 有効なデータのみフィルター
            valid_data = athlete_data.dropna(subset=[metric])
            valid_data = valid_data[valid_data[metric] != 0]
            
            if len(valid_data) < 1:
                continue
            
            # 数値変換
            valid_data[metric] = pd.to_numeric(valid_data[metric], errors='coerce')
            valid_data = valid_data.dropna(subset=[metric])
            
            if len(valid_data) < 1:
                continue
            
            color = athlete_colors[j % len(athlete_colors)]
            
            # ラインプロット
            fig.add_trace(
                go.Scatter(
                    x=valid_data['Date'],
                    y=valid_data[metric],
                    mode='lines+markers',
                    name=athlete,  # 選手名のみを凡例に表示
                    line=dict(
                        color=color,
                        width=3
                    ),
                    marker=dict(
                        size=8,
                        line=dict(width=2, color='white'),
                        symbol='circle'
                    ),
                    legendgroup=athlete,
                    showlegend=(i == 0),  # 最初のメトリクスでのみ凡例表示
                    hovertemplate=f'<b>{athlete}</b><br>Date: %{{x}}<br>{metric}: %{{y:.2f}}<extra></extra>'
                ),
                row=row, col=col
            )
        
        # 軸の設定
        unit = units.get(metric, '')
        fig.update_yaxes(
            title_text=f"{unit}" if unit else "Value",
            row=row, col=col,
            gridcolor='rgba(0,0,0,0.08)',
            linecolor='rgba(0,0,0,0.2)',
            title_font=dict(size=12, color='#2D3748'),
            tickfont=dict(size=10)
        )
        fig.update_xaxes(
            row=row, col=col,
            gridcolor='rgba(0,0,0,0.08)',
            linecolor='rgba(0,0,0,0.2)',
            tickfont=dict(size=10)
        )
    
    fig.update_layout(
        title=dict(
            text=f"{test_config['name']} - Team Comparison",
            x=0.5,
            font=dict(size=20, color='#2D3748', family='Arial Black')
        ),
        height=400 * rows,
        plot_bgcolor='rgba(247, 250, 252, 0.3)',
        paper_bgcolor='white',
        margin=dict(l=50, r=50, t=80, b=50),
        font=dict(family="Arial"),
        legend=dict(
            bgcolor='rgba(255,255,255,0.9)',
            bordercolor='rgba(0,0,0,0.1)',
            borderwidth=1,
            font=dict(size=12)
        )
    )
    
    return fig

def main():
    # Header
    st.markdown('<div class="main-header">Fencing Performance Test</div>', 
                unsafe_allow_html=True)
    
    # Page navigation
    st.markdown('<div class="page-nav">', unsafe_allow_html=True)
    page = st.selectbox(
        "Select Analysis Type",
        ["Individual Analysis", "Team Analysis"],
        help="Choose between individual athlete analysis or team performance trends"
    )
    st.markdown('</div>', unsafe_allow_html=True)
    
    # File upload
    uploaded_file = st.file_uploader(
        "Upload your performance data file",
        type=['xlsx', 'xls'],
        help="Please upload Excel file with 'CMJ' and 'IMTP' sheets"
    )
    
    if uploaded_file is None:
        st.info("Please upload a data file to begin analysis.")
        st.markdown("""
        ### Expected Data Format:
        Excel file with two sheets: 'CMJ' and 'IMTP'
        
        Each sheet should have:
        - Column A: ID
        - Column B: Name
        - Column C: Date
        - Remaining columns: Test metrics
        """)
        st.stop()
    
    # Load data using manual method
    st.info("Loading data...")
    
    try:
        # 手動でExcelを読み込み
        data_dict = load_excel_manually(uploaded_file)
        
        if data_dict is None:
            st.error("Failed to load Excel file")
            st.stop()
        
        # DataFrameを作成
        df = create_dataframe_from_dict(data_dict)
        
        if df.empty:
            st.error("No valid data found")
            st.stop()
        
        # 重複処理
        for test_type in ['CMJ', 'IMTP']:
            test_data = df[df['Type'] == test_type]
            
            if test_type == 'CMJ' and 'Jump Height(cm)' in test_data.columns:
                test_data = test_data.dropna(subset=['Jump Height(cm)'])
                test_data['Jump Height(cm)'] = pd.to_numeric(test_data['Jump Height(cm)'], errors='coerce')
                test_data = test_data.dropna(subset=['Jump Height(cm)'])
                if not test_data.empty:
                    test_data = test_data.sort_values('Jump Height(cm)', ascending=False)
                    test_data = test_data.drop_duplicates(subset=['Name', 'Date'], keep='first')
                    df = df[df['Type'] != test_type]
                    df = pd.concat([df, test_data], ignore_index=True)
            
            elif test_type == 'IMTP' and 'Relative Peak Force (BW)' in test_data.columns:
                test_data = test_data.dropna(subset=['Relative Peak Force (BW)'])
                test_data['Relative Peak Force (BW)'] = pd.to_numeric(test_data['Relative Peak Force (BW)'], errors='coerce')
                test_data = test_data.dropna(subset=['Relative Peak Force (BW)'])
                if not test_data.empty:
                    test_data = test_data.sort_values('Relative Peak Force (BW)', ascending=False)
                    test_data = test_data.drop_duplicates(subset=['Name', 'Date'], keep='first')
                    df = df[df['Type'] != test_type]
                    df = pd.concat([df, test_data], ignore_index=True)
        
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        st.stop()
    
    # Test configuration
    config = get_test_config()
    
    # Individual Analysis Page
    if page == "Individual Analysis":
        # Athlete selection
        available_names = df['Name'].dropna().unique()
        if len(available_names) == 0:
            st.error("No athlete data found.")
            st.stop()
        
        selected_name = st.selectbox("Select Athlete", available_names)
        player_data = df[df['Name'] == selected_name]
        
        if player_data.empty:
            st.error(f"No data found for athlete '{selected_name}'.")
            st.stop()
        
        # Display athlete info
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown(f'<div class="player-title">{selected_name}</div>', unsafe_allow_html=True)
        with col2:
            all_dates = player_data['Date'].dropna()
            if not all_dates.empty:
                # 日付を確実にdatetime型に変換してからソート
                all_dates = pd.to_datetime(all_dates).sort_values(ascending=False)
                latest_date = all_dates.iloc[0].strftime('%Y-%m-%d')
                oldest_date = all_dates.iloc[-1].strftime('%Y-%m-%d')
                st.markdown(f'<div class="date-info">Test Period: {oldest_date} ~ {latest_date}</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="date-info">Test Date: N/A</div>', unsafe_allow_html=True)
        
        # Process each test type
        for test_type, test_config in config.items():
            test_player_data = player_data[player_data['Type'] == test_type]
            
            if test_player_data.empty:
                continue
            
            st.markdown(f'<div class="section-header">{test_config["name"]} ({test_type})</div>', unsafe_allow_html=True)
            
            # Key Indicators
            if test_config['highlight']:
                st.markdown("### Key Indicators")
                highlight_cols = st.columns(len(test_config['highlight']))
                
                for i, metric in enumerate(test_config['highlight']):
                    with highlight_cols[i]:
                        player_val = safe_get_value(test_player_data, metric)
                        best_val, best_date = safe_get_best_value(test_player_data, metric)
                        test_data = df[df['Type'] == test_type]
                        avg_val = safe_mean(test_data[metric])
                        unit = test_config['units'].get(metric, '')
                        
                        female_norm_text = ""
                        if 'female_norms' in test_config and metric in test_config['female_norms']:
                            norm_data = test_config['female_norms'][metric]
                            female_norm_text = f"<br>Female Norm: {norm_data['mean']:.2f} ± {norm_data['std']:.2f}"
                        
                        best_text = ""
                        if best_val is not None:
                            best_text = f"<br>Personal Best: {best_val:.2f}{unit}"
                            if best_date != "N/A":
                                best_text += f" ({best_date})"
                        
                        st.markdown(f"""
                        <div class="metric-card">
                            <div class="metric-label">{metric}</div>
                            <div class="highlight-metric">{format_value(player_val, unit)}</div>
                            <div class="comparison-text">
                                Team Average: {format_value(avg_val, unit)}{best_text}{female_norm_text}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
            
            # Detailed data table
            st.markdown("### Detailed Data")
            available_metrics = [m for m in test_config['metrics'] if m in df.columns]
            
            if available_metrics:
                comparison_df = create_comparison_table(
                    test_player_data, df, available_metrics, test_type, config
                )
                st.dataframe(comparison_df, use_container_width=True, hide_index=True)
                
                # メトリクス選択とトレンドグラフ
                st.markdown("### Progress Chart")
                
                # グラフ表示するメトリクスを選択
                selected_metrics = st.multiselect(
                    f"Select metrics to display for {test_type}",
                    available_metrics,
                    default=available_metrics[:3] if len(available_metrics) >= 3 else available_metrics,
                    key=f"metrics_{test_type}_{selected_name}"
                )
                
                if selected_metrics and PLOTLY_AVAILABLE:
                    try:
                        # サブプロット作成
                        rows = (len(selected_metrics) + 1) // 2
                        cols = min(2, len(selected_metrics))
                        
                        fig = make_subplots(
                            rows=rows,
                            cols=cols,
                            subplot_titles=selected_metrics,
                            vertical_spacing=0.2,
                            horizontal_spacing=0.15
                        )
                        
                        for i, metric in enumerate(selected_metrics):
                            row = (i // 2) + 1
                            col = (i % 2) + 1
                            
                            if metric in test_player_data.columns:
                                # データを準備
                                chart_data = test_player_data[['Date', metric]].dropna()
                                chart_data['Date'] = pd.to_datetime(chart_data['Date'])
                                chart_data = chart_data.sort_values('Date')
                                
                                if not chart_data.empty:
                                    # グラフを追加
                                    mode = 'lines+markers' if len(chart_data) > 1 else 'markers'
                                    fig.add_trace(go.Scatter(
                                        x=chart_data['Date'],
                                        y=chart_data[metric],
                                        mode=mode,
                                        name=metric,
                                        line=dict(color='#2D3748', width=3),
                                        marker=dict(size=8, color='#2D3748'),
                                        showlegend=False
                                    ), row=row, col=col)
                                    
                                    # 軸ラベル設定
                                    unit = test_config['units'].get(metric, '')
                                    fig.update_yaxes(title_text=unit, row=row, col=col)
                                    fig.update_xaxes(title_text="Date", row=row, col=col)
                        
                        # レイアウト設定
                        fig.update_layout(
                            title=f"{test_config['name']} Progress",
                            height=400 * rows,
                            showlegend=False
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                        st.success(f"Chart created successfully for {len(selected_metrics)} metrics!")
                    
                    except Exception as e:
                        st.error(f"Chart creation failed: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
                
                elif not selected_metrics:
                    st.info("Please select at least one metric to display.")
                else:
                    st.error("Plotly not available for chart creation.")
            else:
                st.info(f"No {test_type} data available.")
        
        # レポート生成セクション
        st.markdown("---")
        st.markdown('<div class="report-section">', unsafe_allow_html=True)
        st.markdown("### 📊 Individual Performance Report")
        
        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("""
            **Generate comprehensive A4 report including:**
            - Individual performance summary and trends
            - Team comparison and benchmarks
            - Key metrics: Jump Height, mRSI, Braking RFD, Relative Peak Force
            """)
        
        with col2:
            if st.button("📄 Generate PDF Report", type="primary", use_container_width=True):
                try:
                    with st.spinner("Generating PDF report..."):
                        pdf_data = generate_pdf_report(player_data, df, selected_name)
                    
                    st.download_button(
                        label="📥 Download Report",
                        data=pdf_data,
                        file_name=f"Performance_Report_{selected_name}_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    st.success("✅ Report generated successfully!")
                    
                except Exception as e:
                    st.error(f"Report generation failed: {str(e)}")
                    st.info("Please ensure matplotlib and seaborn are installed")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Team Analysis Page
    elif page == "Team Analysis":
        st.markdown('<div class="section-header">Team Analysis</div>', unsafe_allow_html=True)
        
        # 選手選択
        st.markdown("### Select Athletes for Comparison")
        available_names = df['Name'].dropna().unique()
        selected_athletes = st.multiselect(
            "Choose athletes to compare",
            available_names,
            default=available_names[:3] if len(available_names) >= 3 else available_names,
            help="Select multiple athletes to compare their performance trends"
        )
        
        if selected_athletes:
            st.success(f"Selected {len(selected_athletes)} athletes: {', '.join(selected_athletes)}")
            
            # CMJとIMTPの比較グラフ
            for test_type, test_config in config.items():
                # そのテストタイプのデータが存在するかチェック
                test_data = df[(df['Type'] == test_type) & (df['Name'].isin(selected_athletes))]
                
                if test_data.empty:
                    continue
                
                st.markdown(f'<div class="section-header">{test_config["name"]} ({test_type}) Comparison</div>', unsafe_allow_html=True)
                
                # 比較グラフを作成
                comparison_fig = create_team_comparison_chart(df, selected_athletes, test_type, config)
                
                if comparison_fig:
                    st.plotly_chart(comparison_fig, use_container_width=True, config={'displayModeBar': False})
                else:
                    st.info(f"No sufficient data for {test_type} comparison chart.")
        
        else:
            st.warning("Please select at least one athlete for comparison.")
        
        # 基本チーム統計
        st.markdown("### Team Statistics")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            total_athletes = len(df['Name'].unique())
            st.metric("Total Athletes", total_athletes)
        with col2:
            total_tests = len(df)
            st.metric("Total Tests", total_tests)
        with col3:
            cmj_total = len(df[df['Type'] == 'CMJ'])
            st.metric("CMJ Tests", cmj_total)
        with col4:
            imtp_total = len(df[df['Type'] == 'IMTP'])
            st.metric("IMTP Tests", imtp_total)
        
        # 各テストタイプの統計
        for test_type, test_config in config.items():
            test_data = df[df['Type'] == test_type]
            
            if test_data.empty:
                continue
            
            st.markdown(f"#### {test_config['name']} ({test_type}) Statistics")
            
            # 各メトリクスの統計を計算
            stats_data = []
            for metric in test_config['metrics']:
                if metric in test_data.columns:
                    metric_data = pd.to_numeric(test_data[metric], errors='coerce').dropna()
                    metric_data = metric_data[metric_data != 0]
                    
                    if len(metric_data) > 0:
                        stats_data.append({
                            'Metric': metric,
                            'Count': len(metric_data),
                            'Mean': f"{metric_data.mean():.2f}",
                            'Std Dev': f"{metric_data.std():.2f}",
                            'Min': f"{metric_data.min():.2f}",
                            'Max': f"{metric_data.max():.2f}"
                        })
            
            if stats_data:
                stats_df = pd.DataFrame(stats_data)
                st.dataframe(stats_df, use_container_width=True, hide_index=True)
            else:
                st.info(f"No valid data for {test_type} statistics.")

if __name__ == "__main__":
    main()